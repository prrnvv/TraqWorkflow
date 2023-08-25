from openpyxl import load_workbook
def get_login_data_positive():
    workbook = load_workbook("testdata/testdata.xlsx")
    sheet = workbook["positive_test_data"]
    positive_data = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2)]
    return positive_data

def get_login_data_negative():
    workbook = load_workbook("testdata/testdata.xlsx")
    sheet = workbook["negative_test_data"]
    negative_data = [(row[0].value, row[1].value) for row in sheet.iter_rows(min_row=2)]
    return negative_data